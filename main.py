import customtkinter as ctk
import mysql.connector
from mysql.connector import Error
import requests
from tkinter import ttk, messagebox, filedialog
import json
from datetime import datetime
import os
from openpyxl import Workbook
import math
from tkinter import Canvas, Frame
import tkinter as tk

ctk.set_appearance_mode("dark")


class ScrollableTableGrid(ctk.CTkFrame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.canvas, self.frame = Canvas(
            self, borderwidth=0, highlightthickness=0, width=300, height=40
        ), Frame(self.canvas, width=300, height=40)
        self.vsb = ctk.CTkScrollbar(
            self, orientation="vertical", command=self.canvas.yview
        )
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window(
            (4, 4), window=self.frame, anchor="nw", tags="self.frame"
        )
        self.frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        self.canvas.bind_all(
            "<MouseWheel>",
            lambda e: self.canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"),
        )

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class ConnectionWindow(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Connexion à la base de données")
        self.geometry("400x300")
        self.resizable(False, False)
        self.grab_set()
        self.focus_set()
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)
        ctk.CTkLabel(
            self, text="Connexion MySQL", font=ctk.CTkFont(size=20, weight="bold")
        ).grid(row=0, column=0, pady=(20, 10), sticky="ew")
        self.frame = ctk.CTkFrame(self)
        self.frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.host_entry, self.user_entry, self.password_entry, self.database_entry = [
            self.create_entry(
                self.frame, label, i, show="*" if label == "Mot de passe:" else None
            )
            for i, label in enumerate(
                ["Hôte:", "Utilisateur:", "Mot de passe:", "Base de données:"]
            )
        ]
        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=2, column=0, padx=20, pady=(10, 20), sticky="ew")
        button_frame.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkButton(button_frame, text="Se connecter", command=self.connect).grid(
            row=0, column=0, padx=(0, 10), pady=10, sticky="ew"
        )
        ctk.CTkButton(button_frame, text="Annuler", command=self.destroy).grid(
            row=0, column=1, padx=(10, 0), pady=10, sticky="ew"
        )

    def create_entry(self, parent, label_text, row, show=None):
        ctk.CTkLabel(parent, text=label_text).grid(
            row=row, column=0, padx=(10, 5), pady=5, sticky="e"
        )
        entry = ctk.CTkEntry(parent, show=show)
        entry.grid(row=row, column=1, padx=(5, 10), pady=5, sticky="ew")
        return entry

    def connect(self):
        data = {
            entry: getattr(self, f"{entry}_entry").get()
            for entry in ["host", "user", "password", "database"]
        }
        if not all(data.values()):
            messagebox.showwarning(
                "Champs manquants", "Veuillez remplir tous les champs."
            )
            return
        self.master.connect_to_database(
            data["host"], data["user"], data["password"], data["database"]
        )
        self.destroy()


class AIAssistantWindow(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Assistant IA")
        self.geometry("600x400")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.chat_history = ctk.CTkTextbox(self, state="disabled")
        self.chat_history.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.input_frame = ctk.CTkFrame(self)
        self.input_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        self.input_frame.grid_columnconfigure(0, weight=1)
        self.user_input = ctk.CTkEntry(self.input_frame)
        self.user_input.grid(row=0, column=0, padx=(0, 10), sticky="ew")
        ctk.CTkButton(self.input_frame, text="Envoyer", command=self.send_message).grid(
            row=0, column=1
        )

    def send_message(self):
        user_message = self.user_input.get()
        self.user_input.delete(0, "end")
        self.update_chat_history(f"Vous : {user_message}")
        ai_response = self.get_ai_response(user_message)
        self.update_chat_history(f"Assistant : {ai_response}")

    def update_chat_history(self, message):
        self.chat_history.configure(state="normal")
        self.chat_history.insert("end", message + "\n\n")
        self.chat_history.configure(state="disabled")
        self.chat_history.see("end")

    def get_ai_response(self, user_message):
        try:
            response = requests.post(
                "https://api.mistral.ai/v1/chat/completions",
                headers={
                    "Authorization": "Bearer TAYWvVnNyLn7gyHH3soHMIQjpOXwl9xE",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "mistral-tiny",
                    "messages": [
                        {"role": "system", "content": "Assistant SQL"},
                        {
                            "role": "user",
                            "content": f"Transforme la demande suivante en SQL : '{user_message}'.",
                        },
                    ],
                    "max_tokens": 150,
                    "temperature": 0.3,
                },
            )
            response.raise_for_status()
            return (
                response.json()
                .get("choices", [{}])[0]
                .get("message", {})
                .get("content", "Erreur de réponse.")
                .strip()
            )
        except Exception as e:
            return f"Erreur : {str(e)}"


class CommandsWindow(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Liste des commandes")
        self.geometry("1500x700")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grab_set()
        self.focus_set()
        self.commands = [
            (
                "CREATE TABLE",
                "Crée une nouvelle table",
                "Ex: CREATE TABLE nom_table (id INT, nom VARCHAR(255));",
            ),
            (
                "INSERT INTO",
                "Ajoute une nouvelle ligne",
                "Ex: INSERT INTO nom_table (id, nom) VALUES (1, 'Alice');",
            ),
            ("DROP TABLE", "Supprime une table", "Ex: DROP TABLE nom_table;"),
            (
                "SELECT *",
                "Sélectionne toutes les lignes",
                "Ex: SELECT * FROM nom_table;",
            ),
            (
                "UPDATE",
                "Met à jour une ligne",
                "Ex: UPDATE nom_table SET nom = 'Bob' WHERE id = 1;",
            ),
            (
                "DELETE",
                "Supprime des lignes",
                "Ex: DELETE FROM nom_table WHERE id = 1;",
            ),
            (
                "TRUNCATE TABLE",
                "Supprime toutes les lignes",
                "Ex: TRUNCATE TABLE nom_table;",
            ),
            (
                "CREATE INDEX",
                "Crée un index",
                "Ex: CREATE INDEX idx_nom ON nom_table (nom);",
            ),
            ("DROP INDEX", "Supprime un index", "Ex: DROP INDEX idx_nom ON nom_table;"),
            (
                "ALTER TABLE ADD",
                "Ajoute une colonne",
                "Ex: ALTER TABLE nom_table ADD age INT;",
            ),
            (
                "ALTER TABLE DROP",
                "Supprime une colonne",
                "Ex: ALTER TABLE nom_table DROP COLUMN age;",
            ),
            (
                "RENAME TABLE",
                "Renomme une table",
                "Ex: RENAME TABLE nom_table TO nouvelle_table;",
            ),
            (
                "CREATE USER",
                "Crée un utilisateur",
                "Ex: CREATE USER 'utilisateur'@'localhost' IDENTIFIED BY 'mot_de_passe';",
            ),
            (
                "DROP USER",
                "Supprime un utilisateur",
                "Ex: DROP USER 'utilisateur'@'localhost';",
            ),
            (
                "GRANT",
                "Donne des droits",
                "Ex: GRANT ALL PRIVILEGES ON base.* TO 'utilisateur'@'localhost';",
            ),
            (
                "REVOKE",
                "Révoque des droits",
                "Ex: REVOKE ALL PRIVILEGES ON base.* FROM 'utilisateur'@'localhost';",
            ),
            (
                "BACKUP DATABASE",
                "Sauvegarde une base",
                "Ex: mysqldump -u utilisateur -p base > sauvegarde.sql;",
            ),
            (
                "RESTORE DATABASE",
                "Restaure une base",
                "Ex: mysql -u utilisateur -p base < sauvegarde.sql;",
            ),
            ("BEGIN", "Démarre une transaction", "Ex: BEGIN;"),
            ("COMMIT", "Valide une transaction", "Ex: COMMIT;"),
            ("ROLLBACK", "Annule une transaction", "Ex: ROLLBACK;"),
            (
                "CREATE VIEW",
                "Crée une vue",
                "Ex: CREATE VIEW vue AS SELECT * FROM table;",
            ),
            ("DROP VIEW", "Supprime une vue", "Ex: DROP VIEW vue;"),
            (
                "CREATE PROCEDURE",
                "Crée une procédure",
                "Ex: CREATE PROCEDURE nom_procedure() BEGIN ... END;",
            ),
            (
                "DROP PROCEDURE",
                "Supprime une procédure",
                "Ex: DROP PROCEDURE nom_procedure;",
            ),
            ("EXPLAIN", "Explique une requête", "Ex: EXPLAIN SELECT * FROM nom_table;"),
            ("SHOW TABLES", "Liste les tables", "Ex: SHOW TABLES;"),
        ]
        self.tree = ttk.Treeview(
            self, columns=("Command", "Description", "Example"), show="headings"
        )
        for col, text in zip(
            self.tree["columns"], ["Commande", "Description", "Exemple"]
        ):
            self.tree.heading(col, text=text)
            self.tree.column(col, anchor="w", width=300 if col == "Command" else 450)
        self.tree.grid(row=0, column=0, sticky="nsew")
        for command in self.commands:
            self.tree.insert("", "end", values=command)
        ctk.CTkButton(
            self, text="Télécharger la liste", command=self.download_commands
        ).grid(row=1, column=0, pady=10)
        ctk.CTkButton(self, text="Fermer", command=self.destroy).grid(
            row=2, column=0, pady=10
        )

    def download_commands(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Enregistrer la liste des commandes",
        )
        if file_path:
            self.save_commands_to_excel(file_path)

    def save_commands_to_excel(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste des commandes"
        ws.append(["Commande", "Description", "Exemple"])
        for row in self.tree.get_children():
            ws.append(self.tree.item(row)["values"])
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(file_path)
        messagebox.showinfo(
            "Téléchargement terminé", f"La liste a été enregistrée sous '{file_path}'"
        )


class DatabaseHistory:
    def __init__(self, filename="database_history.json"):
        self.filename = filename
        self.history = self.load_history()

    def load_history(self):
        if os.path.exists(self.filename):
            with open(self.filename, "r") as f:
                return json.load(f)
        return []

    def save_history(self):
        with open(self.filename, "w") as f:
            json.dump(self.history, f)

    def add_entry(self, command, result):
        self.history.append(
            {
                "timestamp": datetime.now().isoformat(),
                "command": command,
                "result": result,
            }
        )
        self.save_history()


class HistoryWindow(ctk.CTkToplevel):
    def __init__(self, master, history):
        super().__init__(master)
        self.title("Historique de la base de données")
        self.geometry("800x600")
        self.history = history
        self.tree = ttk.Treeview(
            self, columns=("Timestamp", "Command", "Result"), show="headings"
        )
        self.tree.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        for col, width in zip(self.tree["columns"], (200, 300, 300)):
            self.tree.heading(col, text=col), self.tree.column(col, width=width)
        ttk.Scrollbar(self, orient="vertical", command=self.tree.yview).grid(
            row=0, column=1, sticky="ns"
        )
        self.load_history()
        ctk.CTkButton(self, text="Télécharger", command=self.download_history).grid(
            row=1, column=0, pady=10
        )

    def load_history(self):
        for entry in self.history.history:
            self.tree.insert(
                "",
                "end",
                values=(entry["timestamp"], entry["command"], entry["result"]),
            )

    def download_history(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if file_path:
            self.save_history_to_excel(file_path)

    def save_history_to_excel(self, file_path):
        wb, ws = Workbook(), Workbook().active
        ws.append(["Horodatage", "Commande", "Résultat"])
        for entry in self.history.history:
            ws.append([entry["timestamp"], entry["command"], entry["result"]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = (
                max(len(str(cell.value)) for cell in col) + 2
            )
        wb.save(file_path)
        messagebox.showinfo(
            "Téléchargement terminé",
            f"L'historique a été téléchargé dans '{file_path}'",
        )


class DatabaseOverviewWindow(ctk.CTkToplevel):
    def __init__(self, master, connection):
        super().__init__(master)
        self.title("Aperçu de la base de données")
        self.geometry("800x600")
        self.connection = connection
        self.tree = ttk.Treeview(
            self, columns=("Table", "Columns", "Rows"), show="headings"
        )
        self.tree.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        for col, width in zip(self.tree["columns"], (200, 400, 100)):
            self.tree.heading(col, text=col), self.tree.column(col, width=width)
        ttk.Scrollbar(self, orient="vertical", command=self.tree.yview).grid(
            row=0, column=1, sticky="ns"
        )
        self.load_database_overview()
        ctk.CTkButton(
            self, text="Télécharger l'aperçu", command=self.download_overview
        ).grid(row=1, column=0, pady=10)

    def load_database_overview(self):
        try:
            with self.connection.cursor() as cursor:
                cursor.execute("SHOW TABLES")
                for table in cursor.fetchall():
                    cursor.execute(f"DESCRIBE {table[0]}")
                    column_names = ", ".join(col[0] for col in cursor.fetchall())
                    cursor.execute(f"SELECT COUNT(*) FROM {table[0]}")
                    row_count = cursor.fetchone()[0]
                    self.tree.insert(
                        "", "end", values=(table[0], column_names, row_count)
                    )
        except Error as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement : {e}")

    def download_overview(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if file_path:
            wb, ws = Workbook(), Workbook().active
            ws.append(["Table", "Colonnes", "Lignes"])
            for item in self.tree.get_children():
                ws.append(self.tree.item(item)["values"])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = (
                    max(len(str(cell.value)) for cell in col) + 2
                )
            wb.save(file_path)
            messagebox.showinfo(
                "Téléchargement terminé",
                f"L'aperçu a été téléchargé dans '{file_path}'",
            )


class MySQLDBAInterface(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.setup_window()
        self.setup_variables()
        self.create_widgets()
        self.setup_grid()

    def setup_window(self):
        self.title("Assistant DBA MySQL")
        self.geometry("1400x800")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

    def setup_variables(self):
        self.current_page = 1
        self.page_size = 100
        self.connection = None
        self.history = DatabaseHistory()
        self.current_table = None
        self.current_filter = ""
        self.current_sort = None
        self.table_buttons = []

    def create_widgets(self):
        self.create_sidebar()
        self.create_main_frame()
        self.create_table_list_frame()

    def setup_grid(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

    def create_main_frame(self):
        self.main_frame = ctk.CTkScrollableFrame(self, fg_color="#2E2E2E")
        self.main_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(4, weight=1)
        self.create_table_selector()
        self.create_table_list_frame()
        self.create_input_area()
        self.create_output_area()
        self.create_result_table()
        self.create_pagination_buttons()

    def create_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(8, weight=1)
        self.logo_label = ctk.CTkLabel(
            self.sidebar, text="MySQL DBA", font=ctk.CTkFont(size=24, weight="bold")
        )
        self.logo_label.grid(row=0, column=0, padx=20, pady=(30, 20))
        buttons = [
            ("Connexion", self.open_connection_window),
            ("Actualiser", self.refresh_table),
            ("Commandes", self.open_commands_window),
            ("Historique", self.open_history_window),
            ("Aperçu BDD", self.open_database_overview),
            ("Assistant IA", self.open_ai_assistant),
        ]
        for i, (text, command) in enumerate(buttons, start=1):
            btn = ctk.CTkButton(
                self.sidebar, text=text, command=command, height=40, corner_radius=8
            )
            btn.grid(row=i, column=0, padx=20, pady=10, sticky="ew")

    def create_table_selector(self):
        self.table_selector_frame = ctk.CTkFrame(self.main_frame, fg_color="#2E2E2E")
        self.table_selector_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.table_selector_frame.grid_columnconfigure(1, weight=1)
        self.search_entry = ctk.CTkEntry(
            self.table_selector_frame,
            placeholder_text="Rechercher une table",
            height=35,
            corner_radius=8,
        )
        self.search_entry.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.search_entry.bind("<KeyRelease>", self.filter_tables)
        self.sort_selector = ctk.CTkComboBox(
            self.table_selector_frame,
            values=[
                "Nom (A-Z)",
                "Nom (Z-A)",
                "ID (croissant)",
                "ID (décroissant)",
                "Date (croissant)",
                "Date (décroissant)",
            ],
            height=35,
            corner_radius=8,
        )
        self.sort_selector.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.sort_selector.set("Trier par")
        self.apply_sort_button = ctk.CTkButton(
            self.table_selector_frame,
            text="Appliquer le tri",
            command=self.apply_sort,
            height=35,
            corner_radius=8,
        )
        self.apply_sort_button.grid(row=0, column=2, padx=10, pady=10)
        self.refresh_button = ctk.CTkButton(
            self.table_selector_frame,
            text="Rafraîchir",
            command=self.refresh_table,
            height=35,
            corner_radius=8,
        )
        self.refresh_button.grid(row=0, column=3, padx=10, pady=10)

    def create_input_area(self):
        self.input_frame = ctk.CTkFrame(self.main_frame)
        self.input_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
        self.input_frame.grid_columnconfigure(1, weight=1)
        self.input_label = ctk.CTkLabel(self.input_frame, text="Commande:")
        self.input_label.grid(row=0, column=0, padx=(10, 5), pady=10)
        self.input_entry = ctk.CTkEntry(self.input_frame, height=35, corner_radius=8)
        self.input_entry.grid(row=0, column=1, padx=(5, 10), pady=10, sticky="ew")
        self.execute_button = ctk.CTkButton(
            self.input_frame,
            text="Exécuter",
            command=self.execute_command,
            height=35,
            corner_radius=8,
        )
        self.execute_button.grid(row=1, column=1, padx=(0, 10), pady=5, sticky="e")

    def create_output_area(self):
        self.output_text = ctk.CTkTextbox(self.main_frame, height=80, corner_radius=8)
        self.output_text.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

    def create_result_table(self):
        self.table_frame = ctk.CTkFrame(self.main_frame, fg_color="#2E2E2E")
        self.table_frame.grid(row=3, column=0, padx=20, pady=20, sticky="nsew")
        self.table_frame.grid_columnconfigure(0, weight=1)
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table = ttk.Treeview(self.table_frame, height=50, style="Custom.Treeview")
        self.table.grid(row=0, column=0, sticky="nsew")
        self.setup_table_style()
        self.table_scrollbar_y = ttk.Scrollbar(
            self.table_frame, orient="vertical", command=self.table.yview
        )
        self.table_scrollbar_y.grid(row=0, column=1, sticky="ns")
        self.table_scrollbar_x = ttk.Scrollbar(
            self.table_frame, orient="horizontal", command=self.table.xview
        )
        self.table_scrollbar_x.grid(row=1, column=0, sticky="ew")
        self.table.configure(
            yscrollcommand=self.table_scrollbar_y.set,
            xscrollcommand=self.table_scrollbar_x.set,
        )
        self.table.column("#0", width=0, stretch=tk.NO)

    def setup_table_style(self):
        style = ttk.Style()
        style.configure(
            "Custom.Treeview",
            background="#F8F8F8",
            foreground="#333333",
            rowheight=25,
            bordercolor="#CCCCCC",
            borderwidth=1,
            highlightthickness=0,
            selectbackground="#E0E0E0",
            selectforeground="#333333",
        )
        style.configure(
            "Custom.Frame",
            background="#2E2E2E",
            bordercolor="#555555",
            borderwidth=2,
        )
        style.configure(
            "Custom.Treeview.Heading",
            background="#D3D3D3",
            foreground="#333333",
            font=("Arial", 12, "bold"),
            anchor="center",
            borderwidth=0,
        )

    def create_pagination_buttons(self):
        self.pagination_frame = ctk.CTkFrame(self.main_frame, fg_color="#2E2E2E")
        self.pagination_frame.grid(row=5, column=0, padx=20, pady=10, sticky="ew")
        self.previous_button = ctk.CTkButton(
            self.pagination_frame,
            text="Page précédente",
            command=self.previous_page,
            height=35,
            corner_radius=8,
        )
        self.previous_button.grid(row=0, column=0, padx=10, pady=10)
        self.next_button = ctk.CTkButton(
            self.pagination_frame,
            text="Page suivante",
            command=self.next_page,
            height=35,
            corner_radius=8,
        )
        self.next_button.grid(row=0, column=1, padx=10, pady=10)
        self.page_label = ctk.CTkLabel(
            self.pagination_frame, text=f"Page {self.current_page}"
        )
        self.page_label.grid(row=0, column=2, padx=10, pady=10)

    def connect_to_database(self, host, user, password, database):
        try:
            self.connection = mysql.connector.connect(
                host=host, user=user, password=password, database=database
            )
            if self.connection.is_connected():
                self.output_text.insert("end", "Connecté à la base de données MySQL.\n")
                self.refresh_table()
        except Error as e:
            self.output_text.insert(
                "end", f"Erreur lors de la connexion à MySQL: {e}\n"
            )
            messagebox.showerror("Erreur de connexion", str(e))

    def execute_command(self):
        if not self.check_connection():
            return
        command = self.input_entry.get()
        try:
            with self.connection.cursor() as cursor:
                cursor.execute(command)
                if cursor.with_rows:
                    result = cursor.fetchall()
                    if result:
                        column_names = [i[0] for i in cursor.description]
                        self.update_table(column_names, result)
                        result_message = "Résultat affiché dans le tableau."
                        self.update_output(result_message)
                        self.download_button.configure(
                            command=lambda: self.download_results(column_names, result)
                        )
                    else:
                        result_message = "Aucun résultat trouvé."
                        self.update_output(result_message)
                else:
                    self.connection.commit()
                    result_message = (
                        "Requête exécutée avec succès (aucun résultat à afficher)."
                    )
                    self.update_output(result_message)
                    self.refresh_table()
                self.history.add_entry(command, result_message)
        except Error as e:
            self.show_error("Erreur d'exécution", str(e))
            self.history.add_entry(command, str(e))

    def check_connection(self):
        if not self.connection or not self.connection.is_connected():
            self.show_error(
                "Non connecté", "Veuillez vous connecter à la base de données d'abord."
            )
            return False
        return True

    def refresh_table(self):
        if not self.check_connection():
            return
        try:
            with self.connection.cursor() as cursor:
                cursor.execute("SHOW TABLES;")
                tables = cursor.fetchall()
                table_names = [table[0] for table in tables]
            self.update_table_list(table_names)
            self.filter_tables(None)
            visible_tables = [
                btn.cget("text") for btn in self.table_buttons if btn.winfo_viewable()
            ]
            if self.current_table and self.current_table in visible_tables:
                self.display_table_content(self.current_table)
            elif visible_tables:
                self.display_table_content(visible_tables[0])
            else:
                self.clear_table()
            if self.current_sort:
                self.apply_sort()

            self.update_output("Tables rafraîchies avec succès.")
        except Error as e:
            self.show_error("Erreur lors du rafraîchissement", str(e))

    def display_table_content(self, table_name):
        if not self.check_connection():
            return
        self.current_table = table_name
        try:
            query = f"SELECT * FROM {table_name} LIMIT {self.page_size} OFFSET {(self.current_page - 1) * self.page_size};"
            with self.connection.cursor() as cursor:
                cursor.execute(query)
                rows = cursor.fetchall()
                columns = [i[0] for i in cursor.description]
            self.update_table(columns, rows)
            self.update_output(
                f"Affichage du contenu de la table '{table_name}', page {self.current_page}."
            )
        except Error as e:
            self.show_error("Erreur lors de l'affichage de la table", str(e))

    def update_table_list(self, table_names):
        for button in self.table_buttons:
            button.destroy()
        self.table_buttons.clear()
        for i, table in enumerate(table_names):
            btn = ctk.CTkButton(
                self.table_list_frame,
                text=table,
                width=60,
                height=25,
                command=lambda t=table: self.display_table_content(t),
            )
            btn.grid(row=i, column=0, padx=2, pady=2, sticky="ew")
            self.table_buttons.append(btn)
        self.table_list_frame.grid_columnconfigure(0, weight=1)

    def filter_tables(self, event):
        self.current_filter = self.search_entry.get().lower()
        for button in self.table_buttons:
            if self.current_filter in button.cget("text").lower():
                button.grid()
            else:
                button.grid_remove()

    def sort_table_content(self, event):
        sort_by = self.sort_selector.get()
        self.current_sort = sort_by
        self.apply_sort()

    def create_table_list_frame(self):
        self.table_list_frame = ctk.CTkScrollableFrame(
            self.main_frame, fg_color="#2E2E2E", height=200
        )
        self.table_list_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.table_list_frame.grid_columnconfigure(0, weight=1)

    def apply_sort(self):
        if not self.current_table:
            self.show_error("Erreur", "Aucune table n'est actuellement sélectionnée.")
            return
        sort_by = self.sort_selector.get()
        if sort_by == "Trier par":
            self.show_error("Erreur", "Veuillez sélectionner un critère de tri.")
            return

        try:
            with self.connection.cursor() as cursor:
                column_name = self.get_sort_column(sort_by)
                order = (
                    "DESC" if "décroissant" in sort_by or "(Z-A)" in sort_by else "ASC"
                )
                query = f"SELECT * FROM {self.current_table} ORDER BY {column_name} {order} LIMIT {self.page_size} OFFSET {(self.current_page - 1) * self.page_size};"
                cursor.execute(query)
                rows = cursor.fetchall()
                columns = [i[0] for i in cursor.description]
            self.update_table(columns, rows)
            self.update_output(f"Tableau trié par {sort_by}")
        except Error as e:
            self.show_error("Erreur lors du tri", str(e))

    def get_sort_column(self, sort_by):
        if "Nom" in sort_by:
            return self.get_first_column_name()
        elif "ID" in sort_by:
            return self.get_id_column_name()
        elif "Date" in sort_by:
            return self.get_date_column_name()
        else:
            raise ValueError("Critère de tri non reconnu")

    def get_first_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            return cursor.fetchone()[0]

    def get_id_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            columns = cursor.fetchall()
            for column in columns:
                if column[3] == "PRI":
                    return column[0]
        return self.get_first_column_name()

    def get_date_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            columns = cursor.fetchall()
            for column in columns:
                if "date" in column[1].lower() or "time" in column[1].lower():
                    return column[0]
        raise ValueError("Aucune colonne de type date trouvée")

    def update_table(self, column_names, rows):
        self.table["columns"] = column_names
        self.table["show"] = "headings"
        for col in column_names:
            self.table.heading(col, text=col)
        self.table.delete(*self.table.get_children())
        for row in rows:
            self.table.insert("", "end", values=row)

    def update_output(self, message):
        self.output_text.insert("end", message + "\n")
        self.output_text.see("end")

    def show_error(self, title, message):
        messagebox.showerror(title, message)

    def open_connection_window(self):
        ConnectionWindow(self)

    def open_commands_window(self):
        CommandsWindow(self)

    def open_history_window(self):
        HistoryWindow(self, self.history)

    def open_database_overview(self):
        if self.check_connection():
            DatabaseOverviewWindow(self, self.connection)

    def open_ai_assistant(self):
        AIAssistantWindow(self)

    def next_page(self):
        self.current_page += 1
        self.page_label.configure(text=f"Page {self.current_page}")
        self.refresh_table_content()

    def previous_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.page_label.configure(text=f"Page {self.current_page}")
            self.refresh_table_content()

    def refresh_table_content(self):
        if hasattr(self, "current_table"):
            self.display_table_content(self.current_table)
        else:
            self.update_output("Veuillez sélectionner une table à afficher.")

    def download_results(self, columns, data):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data:
                sheet.append(row)
            workbook.save(file_path)
            self.update_output("Résultats téléchargés avec succès!")

    def clear_table(self):
        for item in self.table.get_children():
            self.table.delete(item)
        self.table["columns"] = ()
        self.update_output("Aucune table sélectionnée.")


if __name__ == "__main__":
    app = MySQLDBAInterface()
    app.mainloop()
