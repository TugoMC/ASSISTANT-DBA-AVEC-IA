from datetime import datetime
from openpyxl import Workbook
import sys
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLineEdit,
    QTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QLabel,
    QComboBox,
    QDialog,
    QMessageBox,
    QFileDialog,
    QScrollArea,
    QFrame,
)
from PySide6.QtCore import Qt
import mysql.connector
from mysql.connector import Error
import requests


class ScrollableTableGrid(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.table = QTableWidget(self)
        self.table.setRowCount(10)
        self.table.setColumnCount(5)
        self.table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.layout.addWidget(self.table)
        self.table.viewport().installEventFilter(self)

    def eventFilter(self, source, event):
        if event.type() == Qt.Wheel and source == self.table.viewport():
            delta = event.angleDelta().y() / 120
            self.table.verticalScrollBar().setValue(
                self.table.verticalScrollBar().value() - delta * 20
            )
            return True
        return super().eventFilter(source, event)


class ConnectionWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Connexion à la base de données")
        self.setGeometry(300, 300, 300, 200)
        layout = QVBoxLayout()
        self.host_input = QLineEdit()
        self.user_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.database_input = QLineEdit()
        layout.addWidget(QLabel("Hôte:"))
        layout.addWidget(self.host_input)
        layout.addWidget(QLabel("Utilisateur:"))
        layout.addWidget(self.user_input)
        layout.addWidget(QLabel("Mot de passe:"))
        layout.addWidget(self.password_input)
        layout.addWidget(QLabel("Base de données:"))
        layout.addWidget(self.database_input)
        connect_button = QPushButton("Se connecter")
        connect_button.clicked.connect(self.accept)
        layout.addWidget(connect_button)
        self.setLayout(layout)

    def get_connection_info(self):
        return {
            "host": self.host_input.text(),
            "user": self.user_input.text(),
            "password": self.password_input.text(),
            "database": self.database_input.text(),
        }


class AIAssistantWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Assistant IA")
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet(
            """
            QWidget {
                background-color: #2b2b2b;
                color: #ffffff;
                font-family: Arial, sans-serif;
                font-size: 14px;
            }
            QTextEdit, QLineEdit {
                background-color: #3b3b3b;
                border: 1px solid #555555;
                border-radius: 5px;
                padding: 5px;
            }
            QPushButton {
                background-color: #0d6efd;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0b5ed7;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #3b3b3b;
                width: 14px;
                margin: 15px 0 15px 0;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                min-height: 30px;
                border-radius: 7px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #666666;
            }
            QScrollBar::sub-line:vertical, QScrollBar::add-line:vertical {
                border: none;
                background: none;
            }
        """
        )
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        self.chat_history = QTextEdit()
        self.chat_history.setReadOnly(True)
        self.chat_history.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.chat_history.setStyleSheet(
            """
            QTextEdit {
                background-color: #2b2b2b;
                border: none;
                font-size: 15px;
            }
        """
        )
        main_layout.addWidget(self.chat_history)
        input_layout = QHBoxLayout()
        self.input_field = QLineEdit()
        self.input_field.setPlaceholderText("Tapez votre message ici...")
        self.send_button = QPushButton("Envoyer")
        self.send_button.clicked.connect(self.send_message)
        input_layout.addWidget(self.input_field, 7)
        input_layout.addWidget(self.send_button, 1)
        main_layout.addLayout(input_layout)
        self.clear_button = QPushButton("Effacer la conversation")
        self.clear_button.clicked.connect(self.clear_chat)
        main_layout.addWidget(self.clear_button)

    def send_message(self):
        user_message = self.input_field.text()
        if user_message:
            self.add_message("Vous", user_message)
            self.input_field.clear()
            ai_response = self.get_ai_response(user_message)
            self.add_message("Assistant", ai_response)

    def add_message(self, sender, message):
        color = "#4a4a4a" if sender == "Vous" else "#3a3a3a"
        self.chat_history.append(
            f'<div style="background-color: {color}; color: white; margin: 5px; padding: 10px; border-radius: 10px;"><strong>{sender}:</strong> {message}</div>'
        )
        self.chat_history.verticalScrollBar().setValue(
            self.chat_history.verticalScrollBar().maximum()
        )

    def get_ai_response(self, user_message):
        try:
            response = requests.post(
                "https://api.mistral.ai/v1/chat/completions",
                headers={
                    "Authorization": "Bearer TAYWvVnNyLn7gyHH3soHMIQjpOXwl9xE",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "mistral-tiny",
                    "messages": [
                        {"role": "system", "content": "Assistant SQL"},
                        {
                            "role": "user",
                            "content": f"Transforme la demande suivante en SQL : '{user_message}'.",
                        },
                    ],
                    "max_tokens": 150,
                    "temperature": 0.3,
                },
            )
            response.raise_for_status()
            return (
                response.json()
                .get("choices", [{}])[0]
                .get("message", {})
                .get("content", "Erreur de réponse.")
                .strip()
            )
        except Exception as e:
            return f"Erreur : {str(e)}"

    def clear_chat(self):
        self.chat_history.clear()


class CommandsWindow(QMainWindow):
    def __init__(self, master):
        super().__init__(master)
        self.setWindowTitle("Liste des commandes")
        self.setGeometry(100, 100, 1500, 700)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout(self.central_widget)
        self.commands = [
            (
                "CREATE TABLE",
                "Crée une nouvelle table",
                "Ex: CREATE TABLE nom_table (id INT, nom VARCHAR(255));",
            ),
            (
                "INSERT INTO",
                "Ajoute une nouvelle ligne",
                "Ex: INSERT INTO nom_table (id, nom) VALUES (1, 'Alice');",
            ),
            ("DROP TABLE", "Supprime une table", "Ex: DROP TABLE nom_table;"),
            (
                "SELECT *",
                "Sélectionne toutes les lignes",
                "Ex: SELECT * FROM nom_table;",
            ),
            (
                "UPDATE",
                "Met à jour une ligne",
                "Ex: UPDATE nom_table SET nom = 'Bob' WHERE id = 1;",
            ),
            (
                "DELETE",
                "Supprime des lignes",
                "Ex: DELETE FROM nom_table WHERE id = 1;",
            ),
            (
                "TRUNCATE TABLE",
                "Supprime toutes les lignes",
                "Ex: TRUNCATE TABLE nom_table;",
            ),
            (
                "CREATE INDEX",
                "Crée un index",
                "Ex: CREATE INDEX idx_nom ON nom_table (nom);",
            ),
            ("DROP INDEX", "Supprime un index", "Ex: DROP INDEX idx_nom ON nom_table;"),
            (
                "ALTER TABLE ADD",
                "Ajoute une colonne",
                "Ex: ALTER TABLE nom_table ADD age INT;",
            ),
            (
                "ALTER TABLE DROP",
                "Supprime une colonne",
                "Ex: ALTER TABLE nom_table DROP COLUMN age;",
            ),
            (
                "RENAME TABLE",
                "Renomme une table",
                "Ex: RENAME TABLE nom_table TO nouvelle_table;",
            ),
            (
                "CREATE USER",
                "Crée un utilisateur",
                "Ex: CREATE USER 'utilisateur'@'localhost' IDENTIFIED BY 'mot_de_passe';",
            ),
            (
                "DROP USER",
                "Supprime un utilisateur",
                "Ex: DROP USER 'utilisateur'@'localhost';",
            ),
            (
                "GRANT",
                "Donne des droits",
                "Ex: GRANT ALL PRIVILEGES ON base.* TO 'utilisateur'@'localhost';",
            ),
            (
                "REVOKE",
                "Révoque des droits",
                "Ex: REVOKE ALL PRIVILEGES ON base.* FROM 'utilisateur'@'localhost';",
            ),
            (
                "BACKUP DATABASE",
                "Sauvegarde une base",
                "Ex: mysqldump -u utilisateur -p base > sauvegarde.sql;",
            ),
            (
                "RESTORE DATABASE",
                "Restaure une base",
                "Ex: mysql -u utilisateur -p base < sauvegarde.sql;",
            ),
            ("BEGIN", "Démarre une transaction", "Ex: BEGIN;"),
            ("COMMIT", "Valide une transaction", "Ex: COMMIT;"),
            ("ROLLBACK", "Annule une transaction", "Ex: ROLLBACK;"),
            (
                "CREATE VIEW",
                "Crée une vue",
                "Ex: CREATE VIEW vue AS SELECT * FROM table;",
            ),
            ("DROP VIEW", "Supprime une vue", "Ex: DROP VIEW vue;"),
            (
                "CREATE PROCEDURE",
                "Crée une procédure",
                "Ex: CREATE PROCEDURE nom_procedure() BEGIN ... END;",
            ),
            (
                "DROP PROCEDURE",
                "Supprime une procédure",
                "Ex: DROP PROCEDURE nom_procedure;",
            ),
            ("EXPLAIN", "Explique une requête", "Ex: EXPLAIN SELECT * FROM nom_table;"),
            ("SHOW TABLES", "Liste les tables", "Ex: SHOW TABLES;"),
        ]
        self.tree = QTableWidget(self)
        self.tree.setColumnCount(3)
        self.tree.setHorizontalHeaderLabels(["Commande", "Description", "Exemple"])
        self.tree.horizontalHeader().setStretchLastSection(True)
        self.tree.setRowCount(len(self.commands))

        for row, command in enumerate(self.commands):
            for col, value in enumerate(command):
                self.tree.setItem(row, col, QTableWidgetItem(value))

        self.layout.addWidget(self.tree)

        self.download_button = QPushButton("Télécharger la liste", self)
        self.download_button.clicked.connect(self.download_commands)
        self.layout.addWidget(self.download_button)

        self.close_button = QPushButton("Fermer", self)
        self.close_button.clicked.connect(self.close)
        self.layout.addWidget(self.close_button)

    def download_commands(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer la liste des commandes",
            "",
            "Excel files (*.xlsx);;All files (*)",
        )
        if file_path:
            self.save_commands_to_excel(file_path)

    def save_commands_to_excel(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste des commandes"
        ws.append(["Commande", "Description", "Exemple"])
        for row in range(self.tree.rowCount()):
            ws.append(
                [
                    self.tree.item(row, col).text()
                    for col in range(self.tree.columnCount())
                ]
            )
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(file_path)
        QMessageBox.information(
            self,
            "Téléchargement terminé",
            f"La liste a été enregistrée sous '{file_path}'",
        )


class DatabaseHistory:
    def __init__(self):
        self.entries = []

    def add_entry(self, command, result):
        self.entries.append(
            {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "command": command,
                "result": result,
            }
        )


class HistoryWindow(QMainWindow):
    def __init__(self, parent=None, history=None):
        super().__init__(parent)
        self.setWindowTitle("Historique")
        self.setGeometry(100, 100, 800, 600)
        self.history = history
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)
        self.tree = QTableWidget(self)
        self.tree.setColumnCount(3)
        self.tree.setHorizontalHeaderLabels(["Timestamp", "Command", "Result"])
        self.tree.horizontalHeader().setStretchLastSection(True)
        self.layout.addWidget(self.tree)
        self.load_history()
        self.download_button = QPushButton("Télécharger", self)
        self.download_button.clicked.connect(self.download_history)
        self.layout.addWidget(self.download_button)

    def load_history(self):
        if self.history and hasattr(self.history, "entries"):
            self.tree.setRowCount(len(self.history.entries))
            for row, entry in enumerate(self.history.entries):
                self.tree.setItem(row, 0, QTableWidgetItem(entry["timestamp"]))
                self.tree.setItem(row, 1, QTableWidgetItem(entry["command"]))
                self.tree.setItem(row, 2, QTableWidgetItem(entry["result"]))
        else:
            self.tree.setRowCount(0)
            print("No history available or invalid history object")

    def download_history(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer l'historique", "", "Excel files (*.xlsx);;All files (*)"
        )
        if file_path:
            self.save_history_to_excel(file_path)

    def save_history_to_excel(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Historique de la base de données"
        ws.append(["Horodatage", "Commande", "Résultat"])
        for row in range(self.tree.rowCount()):
            ws.append(
                [
                    self.tree.item(row, col).text()
                    for col in range(self.tree.columnCount())
                ]
            )
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(file_path)
        QMessageBox.information(
            self,
            "Téléchargement terminé",
            f"L'historique a été téléchargé dans '{file_path}'",
        )


class DatabaseOverviewWindow(QMainWindow):
    def __init__(self, master=None, connection=None):
        super().__init__(master)
        self.setWindowTitle("Aperçu de la base de données")
        self.setGeometry(100, 100, 800, 600)
        self.connection = connection
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)
        self.tree = QTableWidget(self)
        self.tree.setColumnCount(3)
        self.tree.setHorizontalHeaderLabels(["Table", "Columns", "Rows"])
        self.tree.horizontalHeader().setStretchLastSection(True)
        self.layout.addWidget(self.tree)
        self.load_database_overview()
        self.download_button = QPushButton("Télécharger l'aperçu", self)
        self.download_button.clicked.connect(self.download_overview)
        self.layout.addWidget(self.download_button)

    def load_database_overview(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SHOW TABLES")
            tables = cursor.fetchall()
            self.tree.setRowCount(len(tables))
            for row, table in enumerate(tables):
                cursor.execute(f"DESCRIBE {table[0]}")
                column_names = ", ".join(col[0] for col in cursor.fetchall())
                cursor.execute(f"SELECT COUNT(*) FROM {table[0]}")
                row_count = cursor.fetchone()[0]
                self.tree.setItem(row, 0, QTableWidgetItem(table[0]))
                self.tree.setItem(row, 1, QTableWidgetItem(column_names))
                self.tree.setItem(row, 2, QTableWidgetItem(str(row_count)))
        except Error as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement : {e}")

    def download_overview(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer l'aperçu", "", "Excel files (*.xlsx);;All files (*)"
        )
        if file_path:
            self.save_overview_to_excel(file_path)

    def save_overview_to_excel(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Aperçu de la base de données"
        ws.append(["Table", "Colonnes", "Lignes"])
        for row in range(self.tree.rowCount()):
            ws.append(
                [
                    self.tree.item(row, col).text()
                    for col in range(self.tree.columnCount())
                ]
            )
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(file_path)
        QMessageBox.information(
            self,
            "Téléchargement terminé",
            f"L'aperçu a été téléchargé dans '{file_path}'",
        )


class MySQLDBAInterface(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setup_window()
        self.setup_variables()
        self.create_widgets()

    def setup_window(self):
        self.setWindowTitle("Assistant DBA MySQL")
        self.setGeometry(100, 100, 1400, 800)

    def setup_variables(self):
        self.current_page = 1
        self.page_size = 100
        self.connection = None
        self.history = DatabaseHistory()
        self.current_table = None
        self.current_filter = ""
        self.current_sort = None
        self.table_buttons = []

    def create_widgets(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        self.create_sidebar()
        main_layout.addWidget(self.sidebar)

        self.create_main_frame()
        main_layout.addWidget(self.main_frame)

    def create_sidebar(self):
        self.sidebar = QWidget()
        sidebar_layout = QVBoxLayout(self.sidebar)

        logo_label = QLabel("MySQL DBA")
        logo_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        sidebar_layout.addWidget(logo_label)

        buttons = [
            ("Connexion", self.open_connection_window),
            ("Actualiser", self.refresh_table),
            ("Commandes", self.open_commands_window),
            ("Historique", self.open_history_window),
            ("Aperçu BDD", self.open_database_overview),
            ("Assistant IA", self.open_ai_assistant),
        ]

        for text, command in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(command)
            sidebar_layout.addWidget(btn)

        sidebar_layout.addStretch()

    def create_main_frame(self):
        self.main_frame = QScrollArea()
        self.main_frame.setWidgetResizable(True)
        main_content = QWidget()
        main_layout = QVBoxLayout(main_content)

        self.create_table_selector(main_layout)
        self.create_table_list_frame(main_layout)
        self.create_input_area(main_layout)
        self.create_output_area(main_layout)
        self.create_result_table(main_layout)
        self.create_pagination_buttons(main_layout)

        self.main_frame.setWidget(main_content)

    def create_table_selector(self, parent_layout):
        table_selector_frame = QFrame()
        table_selector_layout = QHBoxLayout(table_selector_frame)

        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Rechercher une table")
        self.search_entry.textChanged.connect(self.filter_tables)
        table_selector_layout.addWidget(self.search_entry)

        self.sort_selector = QComboBox()
        self.sort_selector.addItems(
            [
                "Trier par",
                "Nom (A-Z)",
                "Nom (Z-A)",
                "ID (croissant)",
                "ID (décroissant)",
                "Date (croissant)",
                "Date (décroissant)",
            ]
        )
        table_selector_layout.addWidget(self.sort_selector)

        self.apply_sort_button = QPushButton("Appliquer le tri")
        self.apply_sort_button.clicked.connect(self.apply_sort)
        table_selector_layout.addWidget(self.apply_sort_button)

        self.refresh_button = QPushButton("Rafraîchir")
        self.refresh_button.clicked.connect(self.refresh_table)
        table_selector_layout.addWidget(self.refresh_button)

        parent_layout.addWidget(table_selector_frame)

    def create_table_list_frame(self, parent_layout):
        self.table_list_frame = QScrollArea()
        self.table_list_frame.setWidgetResizable(True)
        table_list_content = QWidget()
        self.table_list_layout = QVBoxLayout(table_list_content)
        self.table_list_frame.setWidget(table_list_content)
        parent_layout.addWidget(self.table_list_frame)

    def create_input_area(self, parent_layout):
        input_frame = QFrame()
        input_layout = QVBoxLayout(input_frame)

        input_label = QLabel("Commande:")
        input_layout.addWidget(input_label)

        self.input_entry = QLineEdit()
        input_layout.addWidget(self.input_entry)

        self.execute_button = QPushButton("Exécuter")
        self.execute_button.clicked.connect(self.execute_command)
        input_layout.addWidget(self.execute_button)

        parent_layout.addWidget(input_frame)

    def create_output_area(self, parent_layout):
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        parent_layout.addWidget(self.output_text)

    def create_result_table(self, parent_layout):
        self.table = QTableWidget()
        parent_layout.addWidget(self.table)

    def create_pagination_buttons(self, parent_layout):
        pagination_frame = QFrame()
        pagination_layout = QHBoxLayout(pagination_frame)

        self.previous_button = QPushButton("Page précédente")
        self.previous_button.clicked.connect(self.previous_page)
        pagination_layout.addWidget(self.previous_button)

        self.next_button = QPushButton("Page suivante")
        self.next_button.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.next_button)

        self.page_label = QLabel(f"Page {self.current_page}")
        pagination_layout.addWidget(self.page_label)

        parent_layout.addWidget(pagination_frame)

    def connect_to_database(self, host, user, password, database):
        try:
            self.connection = mysql.connector.connect(
                host=host, user=user, password=password, database=database
            )
            if self.connection.is_connected():
                self.output_text.append("Connecté à la base de données MySQL.")
                self.refresh_table()
        except Error as e:
            self.output_text.append(f"Erreur lors de la connexion à MySQL: {e}")
            QMessageBox.critical(self, "Erreur de connexion", str(e))

    def execute_command(self):
        if not self.check_connection():
            return
        command = self.input_entry.text()
        try:
            with self.connection.cursor() as cursor:
                cursor.execute(command)
                if (
                    cursor.description
                ):  # Correction : Vérification de la présence de résultats
                    result = cursor.fetchall()
                    if result:
                        column_names = [i[0] for i in cursor.description]
                        self.update_table(column_names, result)
                        result_message = "Résultat affiché dans le tableau."
                        self.update_output(result_message)
                    else:
                        result_message = "Aucun résultat trouvé."
                        self.update_output(result_message)
                else:
                    self.connection.commit()
                    result_message = (
                        "Requête exécutée avec succès (aucun résultat à afficher)."
                    )
                    self.update_output(result_message)
                    self.refresh_table()
                self.history.add_entry(command, result_message)
        except Error as e:
            self.show_error("Erreur d'exécution", str(e))
            self.history.add_entry(command, str(e))

    def check_connection(self):
        if not self.connection or not self.connection.is_connected():
            self.show_error(
                "Non connecté", "Veuillez vous connecter à la base de données d'abord."
            )
            return False
        return True

    def refresh_table(self):
        if not self.check_connection():
            return
        try:
            with self.connection.cursor() as cursor:
                cursor.execute("SHOW TABLES;")
                tables = cursor.fetchall()
                table_names = [table[0] for table in tables]
            self.update_table_list(table_names)
            self.filter_tables()
            visible_tables = [
                btn.text() for btn in self.table_buttons if not btn.isHidden()
            ]
            if self.current_table and self.current_table in visible_tables:
                self.display_table_content(self.current_table)
            elif visible_tables:
                self.display_table_content(visible_tables[0])
            else:
                self.clear_table()
            if self.current_sort:
                self.apply_sort()

            self.update_output("Tables rafraîchies avec succès.")
        except Error as e:
            self.show_error("Erreur lors du rafraîchissement", str(e))

    def display_table_content(self, table_name):
        if not self.check_connection():
            return
        self.current_table = table_name
        try:
            query = f"SELECT * FROM {table_name} LIMIT {self.page_size} OFFSET {(self.current_page - 1) * self.page_size};"
            with self.connection.cursor() as cursor:
                cursor.execute(query)
                rows = cursor.fetchall()
                columns = [i[0] for i in cursor.description]
            self.update_table(columns, rows)
            self.update_output(
                f"Affichage du contenu de la table '{table_name}', page {self.current_page}."
            )
        except Error as e:
            self.show_error("Erreur lors de l'affichage de la table", str(e))

    def update_table_list(self, table_names):
        for button in self.table_buttons:
            self.table_list_layout.removeWidget(button)
            button.deleteLater()
        self.table_buttons.clear()
        for table in table_names:
            btn = QPushButton(table)
            btn.clicked.connect(lambda checked, t=table: self.display_table_content(t))
            self.table_list_layout.addWidget(btn)
            self.table_buttons.append(btn)

    def filter_tables(self):
        self.current_filter = self.search_entry.text().lower()
        for button in self.table_buttons:
            button.setVisible(self.current_filter in button.text().lower())

    def apply_sort(self):
        if not self.current_table:
            self.show_error("Erreur", "Aucune table n'est actuellement sélectionnée.")
            return
        sort_by = self.sort_selector.currentText()
        if sort_by == "Trier par":
            self.show_error("Erreur", "Veuillez sélectionner un critère de tri.")
            return

        try:
            with self.connection.cursor() as cursor:
                column_name = self.get_sort_column(sort_by)
                order = (
                    "DESC" if "décroissant" in sort_by or "(Z-A)" in sort_by else "ASC"
                )
                query = f"SELECT * FROM `{self.current_table}` ORDER BY `{column_name}` {order} LIMIT {self.page_size} OFFSET {(self.current_page - 1) * self.page_size};"
                cursor.execute(query)
                rows = cursor.fetchall()
                columns = [i[0] for i in cursor.description]
            self.update_table(columns, rows)
            self.update_output(f"Tableau trié par {sort_by}")
        except Error as e:
            self.show_error("Erreur lors du tri", str(e))

    def get_sort_column(self, sort_by):
        if "Nom" in sort_by:
            return self.get_first_column_name()
        elif "ID" in sort_by:
            return self.get_id_column_name()
        elif "Date" in sort_by:
            return self.get_date_column_name()
        else:
            raise ValueError("Critère de tri non reconnu")

    def get_first_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            return cursor.fetchone()[0]

    def get_id_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            columns = cursor.fetchall()
            for column in columns:
                if column[3] == "PRI":
                    return column[0]
        return self.get_first_column_name()

    def get_date_column_name(self):
        with self.connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {self.current_table}")
            columns = cursor.fetchall()
            for column in columns:
                if "date" in column[1].lower() or "time" in column[1].lower():
                    return column[0]
        raise ValueError("Aucune colonne de type date trouvée")

    def update_table(self, column_names, rows):
        self.table.setColumnCount(len(column_names))
        self.table.setHorizontalHeaderLabels(column_names)
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(str(value)))

    def update_output(self, message):
        self.output_text.append(message)

    def show_error(self, title, message):
        QMessageBox.critical(self, title, message)

    def open_connection_window(self):
        dialog = ConnectionWindow(self)
        if (
            dialog.exec() == QDialog.Accepted
        ):  # Notez l'utilisation de exec() au lieu de exec_()
            try:
                connection_info = dialog.get_connection_info()
                self.connect_to_database(**connection_info)
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Erreur de connexion",
                    f"Impossible de se connecter à la base de données : {str(e)}",
                )

    def open_commands_window(self):
        commands_window = CommandsWindow(self)
        commands_window.show()

    def open_history_window(self):
        history_window = HistoryWindow(self, self.history)
        history_window.show()

    def open_database_overview(self):
        if self.check_connection():
            overview_window = DatabaseOverviewWindow(self, self.connection)
            overview_window.show()

    def open_ai_assistant(self):
        ai_assistant_window = AIAssistantWindow(self)
        ai_assistant_window.show()

    def next_page(self):
        self.current_page += 1
        self.page_label.setText(f"Page {self.current_page}")
        self.refresh_table_content()

    def previous_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.page_label.setText(f"Page {self.current_page}")
            self.refresh_table_content()

    def refresh_table_content(self):
        if hasattr(self, "current_table"):
            self.display_table_content(self.current_table)
        else:
            self.update_output("Veuillez sélectionner une table à afficher.")

    def download_results(self, columns, data):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Sauvegarder les résultats", "", "Excel files (*.xlsx);;All Files (*)"
        )
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data:
                sheet.append(row)
            workbook.save(file_path)
            self.update_output("Résultats téléchargés avec succès!")

    def clear_table(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.update_output("Aucune table sélectionnée.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MySQLDBAInterface()
    window.show()
    sys.exit(app.exec())
